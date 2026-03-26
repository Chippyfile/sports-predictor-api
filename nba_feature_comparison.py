#!/usr/bin/env python3
"""
nba_feature_comparison.py — Fetch today's games and build feature comparison xlsx.

Usage:
    python3 nba_feature_comparison.py 401810920 401810919 401810918
"""
import sys, json, requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAILWAY_URL = "https://sports-predictor-api-production.up.railway.app"

# The 69 backward-eliminated features
FEATURES_69 = [
    "after_loss_either", "altitude_factor", "ato_ratio_diff",
    "away_is_public_team", "b2b_diff", "bimodal_diff", "ceiling_diff",
    "conference_game", "consistency_diff", "days_since_loss_diff",
    "efg_diff", "elo_diff", "espn_pregame_wp", "espn_pregame_wp_pbp",
    "floor_diff", "ftpct_diff", "games_diff", "games_last_14_diff",
    "h2h_avg_margin", "h2h_total_games", "home_b2b", "home_fav",
    "implied_prob_home", "is_early_season", "is_friday_sat",
    "is_revenge_home", "lineup_value_diff", "margin_accel_diff",
    "market_spread", "matchup_efg", "matchup_ft", "matchup_orb",
    "momentum_halflife_diff", "opp_suppression_diff", "ou_gap",
    "overround", "pace_control_diff", "pace_leverage", "post_allstar",
    "post_trade_deadline", "pyth_luck_diff", "pyth_residual_diff",
    "recovery_diff", "ref_foul_proxy", "ref_home_whistle", "ref_ou_bias",
    "ref_pace_impact", "reverse_line_movement", "roll_bench_pts_diff",
    "roll_ft_trip_rate_diff", "roll_max_run_avg", "roll_paint_fg_rate_diff",
    "roll_paint_pts_diff", "roll_q4_diff", "roll_three_fg_rate_diff",
    "score_kurtosis_diff", "scoring_entropy_diff", "scoring_hhi_diff",
    "sharp_spread_signal", "spread_juice_imbalance", "steals_to_diff",
    "three_pt_regression_diff", "three_value_diff", "threepct_diff",
    "ts_regression_diff", "turnovers_diff", "vig_uncertainty",
    "win_aging_diff", "win_pct_diff",
]

# Feature categories for grouping
CATEGORIES = {
    "Market": ["market_spread", "implied_prob_home", "overround", "home_fav",
               "ou_gap", "sharp_spread_signal", "spread_juice_imbalance",
               "vig_uncertainty", "reverse_line_movement"],
    "ESPN": ["espn_pregame_wp", "espn_pregame_wp_pbp"],
    "Team Stats": ["efg_diff", "ftpct_diff", "threepct_diff", "turnovers_diff",
                    "ato_ratio_diff", "steals_to_diff", "win_pct_diff",
                    "opp_suppression_diff", "three_value_diff", "three_pt_regression_diff",
                    "ts_regression_diff"],
    "Player/Lineup": ["lineup_value_diff", "scoring_hhi_diff", "scoring_entropy_diff"],
    "Enrichment": ["ceiling_diff", "floor_diff", "consistency_diff", "bimodal_diff",
                    "score_kurtosis_diff", "margin_accel_diff", "momentum_halflife_diff",
                    "win_aging_diff", "pyth_residual_diff", "pyth_luck_diff",
                    "pace_control_diff", "pace_leverage", "recovery_diff",
                    "matchup_efg", "matchup_ft", "matchup_orb"],
    "Rolling PBP": ["roll_q4_diff", "roll_paint_pts_diff", "roll_bench_pts_diff",
                     "roll_ft_trip_rate_diff", "roll_three_fg_rate_diff",
                     "roll_paint_fg_rate_diff", "roll_max_run_avg"],
    "Schedule": ["b2b_diff", "home_b2b", "games_last_14_diff", "games_diff",
                  "days_since_loss_diff", "is_early_season", "is_friday_sat",
                  "post_allstar", "post_trade_deadline", "altitude_factor"],
    "H2H": ["h2h_total_games", "h2h_avg_margin", "is_revenge_home",
             "conference_game"],
    "Situational": ["after_loss_either", "away_is_public_team"],
    "Elo": ["elo_diff"],
    "Referee": ["ref_home_whistle", "ref_foul_proxy", "ref_ou_bias", "ref_pace_impact"],
}


def fetch_game(game_id):
    """Fetch prediction from Railway."""
    print(f"  Fetching game {game_id}...")
    try:
        r = requests.post(
            f"{RAILWAY_URL}/predict/nba/full",
            json={"game_id": str(game_id)},
            timeout=30
        )
        if r.ok:
            data = r.json()
            print(f"    {data.get('home_team', '?')} vs {data.get('away_team', '?')} — "
                  f"coverage {data.get('feature_coverage', '?')}")
            return data
        else:
            print(f"    ERROR: {r.status_code} {r.text[:100]}")
            return None
    except Exception as e:
        print(f"    ERROR: {e}")
        return None


def extract_feature_values(prediction):
    """Extract feature values from SHAP output."""
    values = {}
    for s in prediction.get("shap", []):
        values[s["feature"]] = s["value"]
    return values


def build_spreadsheet(games, output_path="nba_feature_comparison.xlsx"):
    """Build the comparison spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Feature Comparison"

    # Colors
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    cat_fill = PatternFill("solid", fgColor="D6E4F0")
    cat_font = Font(bold=True, size=10, name="Arial", color="1F4E79")
    data_font = Font(size=10, name="Arial")
    zero_font = Font(size=10, name="Arial", color="999999")
    nonzero_font = Font(size=10, name="Arial", color="000000")
    positive_fill = PatternFill("solid", fgColor="E2EFDA")
    negative_fill = PatternFill("solid", fgColor="FCE4EC")
    border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )

    # Extract data
    game_data = []
    for g in games:
        if g is None:
            continue
        home = g.get("home_team", "?")
        away = g.get("away_team", "?")
        vals = extract_feature_values(g)
        game_data.append({
            "home": home, "away": away,
            "game_id": g.get("game_id", ""),
            "spread": g.get("market_spread", 0),
            "margin": g.get("ml_margin", 0),
            "wp": g.get("ml_win_prob_home", 0),
            "values": vals,
        })

    n_games = len(game_data)
    if n_games == 0:
        print("  No game data to write!")
        return

    # ── HEADER ROW ──
    ws.cell(row=1, column=1, value="Feature").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    col = 2
    for gd in game_data:
        label = f"{gd['home']} vs {gd['away']}"
        ws.cell(row=1, column=col, value=label).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        col += 1

    # ── SUMMARY ROWS ──
    summary_items = [
        ("Game ID", "game_id"),
        ("Market Spread", "spread"),
        ("ML Predicted Margin", "margin"),
        ("ML Win Prob Home", "wp"),
    ]
    row = 2
    for label, key in summary_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10, name="Arial")
        col = 2
        for gd in game_data:
            val = gd.get(key, "")
            if isinstance(val, float):
                val = round(val, 4)
            ws.cell(row=row, column=col, value=val).font = data_font
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            col += 1
        row += 1

    row += 1  # blank row

    # ── FEATURES BY CATEGORY ──
    for cat_name, cat_features in CATEGORIES.items():
        # Category header
        ws.cell(row=row, column=1, value=cat_name.upper()).font = cat_font
        ws.cell(row=row, column=1).fill = cat_fill
        for c in range(2, 2 + n_games):
            ws.cell(row=row, column=c).fill = cat_fill
        row += 1

        for feat in cat_features:
            if feat not in FEATURES_69:
                continue
            ws.cell(row=row, column=1, value=feat).font = data_font
            ws.cell(row=row, column=1).border = border

            col = 2
            for gd in game_data:
                val = gd["values"].get(feat, 0)
                if val is None:
                    val = 0
                ws.cell(row=row, column=col, value=round(val, 4) if isinstance(val, float) else val)
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=col).border = border

                # Color coding
                if val == 0:
                    ws.cell(row=row, column=col).font = zero_font
                elif isinstance(val, (int, float)):
                    ws.cell(row=row, column=col).font = nonzero_font
                    if val > 0.01:
                        ws.cell(row=row, column=col).fill = positive_fill
                    elif val < -0.01:
                        ws.cell(row=row, column=col).fill = negative_fill
                col += 1
            row += 1
        row += 1  # gap between categories

    # Column widths
    ws.column_dimensions["A"].width = 35
    for c in range(2, 2 + n_games):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # Freeze panes
    ws.freeze_panes = "B7"

    wb.save(output_path)
    print(f"\n  Saved to {output_path}")


if __name__ == "__main__":
    game_ids = sys.argv[1:] if len(sys.argv) > 1 else ["401810920", "401810919", "401810918"]

    print("=" * 60)
    print("  NBA Feature Comparison — Today's Games")
    print("=" * 60)

    games = []
    for gid in game_ids:
        g = fetch_game(gid)
        games.append(g)

    build_spreadsheet(games)
